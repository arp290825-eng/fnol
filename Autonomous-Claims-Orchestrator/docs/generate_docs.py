"""
Generate comprehensive PDF and XLSX documentation for the
Autonomous Claims Orchestrator end-to-end workflow.
All content is derived from actual source files – nothing hallucinated.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────
# DATA  (sourced 100 % from codebase exploration)
# ─────────────────────────────────────────────────────────

WORKFLOW_STEPS = [
    {
        "step": 1,
        "phase": "Email Ingestion",
        "service": "email_ingestion",
        "file": "backend/email_ingestion/service.py",
        "tech": "Python stdlib · imaplib · ssl · email · IMAP4_SSL",
        "llm": "gpt-4o-mini (FNOL classifier, yes/no, max_tokens=10, temp=0)",
        "summary": "Connect to IMAP server and retrieve emails",
        "details": [
            "Load env vars via _load_env() from ENV_FILE (.env)",
            "Open IMAP4_SSL connection (default: imap.gmail.com:993, SSL verify OFF)",
            "Select mailbox: IMAP_MAILBOX env (default INBOX)",
            "SEARCH ALL (if IMAP_SYNC_INCLUDE_READ=true) or UNSEEN",
            "Gmail fallback: tries [Gmail]/All Mail → [Google Mail]/All Mail → INBOX",
            "Fetch each email as RFC822 raw bytes",
            "Parse MIME: extract subject, from, to, date, Message-ID, plain body, attachments",
            "Body extraction: prefers text/plain; fallback text/html (HTML stripped via _strip_html)",
        ],
        "filters": [
            "Duplicate check: Message-ID OR subject|from|date hash",
            "Thread-reply fast-path: if reply to existing claim → merge, skip FNOL gate",
            "_should_reject_consumer_complaint_correspondence: OEM/retail complaint patterns, short replies on quoted threads",
            "_is_clearly_procedural_faq_question: 'how do I file', policy expiry questions → route to FAQ, not FNOL",
            "_has_relevant_keywords / _has_strong_keywords: applied on latest non-quoted segment (_primary_message_text)",
            "LLM FNOL gate (gpt-4o-mini): system='You classify P&C emails – answer only yes or no. Is this FNOL?'; bypass if FNOL_FILTER_ENABLED=false or strong-keyword path + no API key",
        ],
        "policy_extraction": "Regex in ingested_claims/service.py extract_policy_number(); patterns: keyword-anchored (poli?cy\\s*(?:number|no|#...)\\s*[:#-=]\\s*([A-Z0-9]{4,}...)), structural (\\b[A-Z]{2}\\d{6,}\\b), claim-number fallback; if none found: sender email / Message-ID display used as fallback",
        "output": "Saves to data/ingested-claims.json via save_ingested_claim(); attachments stored in data/ingested-attachments/{claimId}/; triggers send_fnol_received_acknowledgement_email() (SMTP) if FNOL_RECEIPT_ACK_EMAIL_ENABLED=true",
    },
    {
        "step": 2,
        "phase": "SendGrid Inbound (Alternate Ingestion)",
        "service": "frontend webhook + ingested_claims",
        "file": "frontend/app/api/webhooks/sendgrid-inbound/route.ts · backend/ingested_claims/service.py",
        "tech": "Next.js API Route · SendGrid Inbound Parse · multipart/form-data",
        "llm": "None",
        "summary": "Receive inbound emails via SendGrid HTTP webhook (alternative to IMAP)",
        "details": [
            "SendGrid posts multipart form-data to /api/webhooks/sendgrid-inbound",
            "Next.js route parses fields: from, to, subject, text, html, attachments",
            "If body is empty, extract_plain_body_from_rfc822() parses raw email field",
            "Calls POST /api/ingested-claims on FastAPI (Python backend) with same payload shape as IMAP path",
            "Python backend deduplicates and stores via same save_ingested_claim() function",
        ],
        "filters": "Same dedup + LLM FNOL gate as IMAP path via shared ingested_claims/service.py logic",
        "policy_extraction": "Same extract_policy_number() regex as IMAP path",
        "output": "Same JSON storage + FNOL ack email as IMAP path",
    },
    {
        "step": 3,
        "phase": "Ingested Claims Storage",
        "service": "ingested_claims",
        "file": "backend/ingested_claims/service.py · mail_chain.py",
        "tech": "Python · JSON file storage · regex · stdlib",
        "llm": "None",
        "summary": "Persist, deduplicate, and thread-merge ingested emails",
        "details": [
            "save_ingested_claim(): writes/updates data/ingested-claims.json",
            "Thread merging: mail_chain.py groups email chains by subject / policy number",
            "Dedup: is_duplicate_email() checks Message-ID first, then subject|from|date fingerprint",
            "get_ingested_claim_by_id(): lookup by claimId",
            "get_all_ingested_claims(): list with optional source filter (faq / fnol)",
            "clear_all_ingested_claims(): wipe for dev/demo reset",
            "extract_policy_number() applied at save time; stored in claim.policyNumber",
        ],
        "filters": "N/A – storage layer",
        "policy_extraction": "Stored at ingest time from regex extraction",
        "output": "data/ingested-claims.json, data/ingested-attachments/{claimId}/*",
    },
    {
        "step": 4,
        "phase": "Claim Extraction",
        "service": "extraction",
        "file": "backend/extraction/service.py",
        "tech": "Python · OpenAI Python SDK · base64 · mimetypes",
        "llm": "gpt-4o (email extraction, document extraction, vision) · temp=0.1 · max_tokens=1500 (email/doc) / 2500 (vision)",
        "summary": "LLM-powered structured field extraction from email body and attachments",
        "details": [
            "Entry: extract_claim_information(claim_id, email_body, attachments)",
            "Step A – Email extraction: extract_from_email(email_body) → JSON schema prompt",
            "  System: 'You are an expert insurance claims processor. Extract JSON only.'",
            "  Fields: policyNumber, claimantName, contactEmail, contactPhone, lossDate,",
            "          lossType (fire/water/auto/theft/liability/medical/other),",
            "          lossLocation, description, vehicleInfo, propertyAddress,",
            "          estimatedAmount, _confidence (per-field 0.0–1.0)",
            "Step B – Attachment processing (per attachment):",
            "  Image files → analyze_image_with_vision(): multimodal message with image_url data URL",
            "    Prompt: structured KPIs (damage severity, repair cost, injury indicators) + detailed_summary",
            "  Text/PDF → extract_from_document(): classify type + extract keyFields",
            "    Document types: PoliceReport, RepairEstimate, Invoice, MedicalRecord,",
            "                    IncidentReport, DamagePhoto, Other",
            "  Binary/unknown → placeholder entry",
            "Step C – Evidence list: builds evidence[] array from scalar field values + confidence scores",
            "Returns: { extractedFields, evidence, documentAnalysis[] }",
        ],
        "filters": "N/A",
        "policy_extraction": "LLM extracts policyNumber from email body (re-extraction with full context, higher accuracy than regex)",
        "output": "Dict: extractedFields{}, evidence[], documentAnalysis[]; passed directly to build_decision_pack()",
    },
    {
        "step": 5,
        "phase": "Policy Grounding (Local DB)",
        "service": "decision / policy_grounding_local",
        "file": "backend/decision/policy_grounding_local.py · database/local_data/*.json",
        "tech": "Python · JSON file DB · regex · rule-based scoring",
        "llm": "None",
        "summary": "Match extracted claim against local customer/policy database to ground coverage",
        "details": [
            "Input: policyNumber (from LLM extraction) OR sender email",
            "find_customer_by_email(): email → customer record from customers.json",
            "find_policy_by_number(): policyNumber → policy record from policies.json",
            "Policy active check: status == 'active' + expiry date >= today",
            "Coverage code lookup: policy_grounding_mapping.json maps lossType × policy_type → covered clause IDs",
            "policy_type detection: AUTO / HOME / COMMERCIAL from policy.policy_type field",
            "Per-clause confidence scoring formula:",
            "  base = 1.0; deduct for inactive policy, expiry within 30 days, lossType mismatch;",
            "  sort descending, return top clauses",
            "build_policy_holder_info(): assembles name, address, contact from customer + policy records (PII masked in draft)",
            "Returns: { groundingClauses[], coverageConfirmed, policyStatus, policyHolderInfo }",
        ],
        "filters": "N/A",
        "policy_extraction": "Resolves policyNumber → full policy record",
        "output": "grounding result dict; if empty → fallback to policy_clauses.py similarity engine",
    },
    {
        "step": 6,
        "phase": "Policy Grounding (ISO Clause Fallback)",
        "service": "decision / policy_clauses",
        "file": "backend/decision/policy_clauses.py",
        "tech": "Python · cosine/keyword similarity · static ISO clause library",
        "llm": "None",
        "summary": "Fallback: similarity-match claim against static ISO policy clauses when local DB has no match",
        "details": [
            "POLICY_CLAUSES: static list of ISO PAP / HO-3 / CGL-style clause texts",
            "_infer_product_types(): from policy number prefix (AC→auto, HO→home, CL→commercial)",
            "_infer_loss_types(): regex on lossType + description text",
            "_compute_similarity(): keyword overlap between clause text and claim description",
            "Threshold: CONFIDENCE_THRESHOLD_HIGH=0.8, MEDIUM=0.6",
            "Only clauses with similarity >= 0.6 returned; max 6 results",
            "get_policy_grounding(): wraps above steps, returns same shape as local grounding",
        ],
        "filters": "similarity < 0.6 → excluded",
        "policy_extraction": "N/A – uses already-extracted policy data",
        "output": "groundingClauses[] with similarity scores; used as fallback in build_decision_pack()",
    },
    {
        "step": 7,
        "phase": "Decision Pack Assembly",
        "service": "decision",
        "file": "backend/decision/service.py",
        "tech": "Python · rule-based logic · Pydantic-style dicts",
        "llm": "None (extraction already ran LLM; decision is deterministic)",
        "summary": "Assemble full structured decision pack combining extraction + grounding results",
        "details": [
            "build_decision_pack(claim_id, extraction_result)",
            "Merge extractedFields from email extraction",
            "Resolve policy: use LLM-extracted policyNumber OR fall back via sender email lookup",
            "Run get_policy_grounding_from_local_data(); if empty → get_policy_grounding() (ISO clauses)",
            "Derive coverage_confirmed: True if grounding returned clauses + policy is active",
            "Policy status flags: POLICY-EXPIRED / POLICY-INACTIVE / POLICY-NOT-FOUND / CUSTOMER-NOT-FOUND",
            "Build policy_holder_info from local DB (name, address, policy limits, deductible)",
            "Assemble decisionPack:",
            "  - claimId, timestamp, processingStatus",
            "  - extractedFields (with _confidence map)",
            "  - evidence[] list",
            "  - documentAnalysis[] (attachment results)",
            "  - groundingClauses[] (applicable policy clauses)",
            "  - coverageConfirmed bool",
            "  - policyHolderInfo",
            "  - auditMetadata (model used, processing time, data sources)",
        ],
        "filters": "Policy status → auto-reject flags if EXPIRED/INACTIVE/NOT-FOUND",
        "policy_extraction": "Final resolution: LLM extraction → DB lookup → grounding",
        "output": "decisionPack dict; consumed by process_claim orchestrator",
    },
    {
        "step": 8,
        "phase": "Process Claim Orchestration",
        "service": "process_claim",
        "file": "backend/process_claim/orchestrator.py",
        "tech": "Python · rule-based orchestration",
        "llm": "Indirect: delegates to extraction (gpt-4o) and FAQ (gpt-4o-mini)",
        "summary": "E2E pipeline coordinator: load → extract → decide → save/reject",
        "details": [
            "POST /api/process-claim → FastAPI → process_claim(claim_id)",
            "1. get_ingested_claim_by_id(claim_id) – load raw ingested claim",
            "2. extract_claim_information() – run extraction pipeline",
            "3. build_decision_pack() – run grounding + decision assembly",
            "4. Auto-reject check: if policyStatus in [POLICY-EXPIRED, POLICY-INACTIVE, POLICY-NOT-FOUND, CUSTOMER-NOT-FOUND]",
            "   AND policyHolderInfo.status == EXPIRED → send_desk_rejection_email()",
            "   → status = 'desk_rejected' → save_processed_claim()",
            "5. Otherwise: status = 'under_review' → save_processed_claim()",
            "   → triggers send_claim_under_review_email() from dashboard/service.py",
        ],
        "filters": "Auto-reject on invalid/expired policy conditions",
        "policy_extraction": "Uses final resolved policyNumber from step 7",
        "output": "Processed claim JSON saved; email notification sent to claimant",
    },
    {
        "step": 9,
        "phase": "Dashboard & Processed Claims Storage",
        "service": "dashboard",
        "file": "backend/dashboard/service.py",
        "tech": "Python · JSON · CSV · file I/O",
        "llm": "None",
        "summary": "Persist processed claims and compute KPIs for the dashboard",
        "details": [
            "save_processed_claim(): write data/processed-claims/{claimId}.json",
            "Update data/processed-claims/claims-index.json",
            "Append row to data/processed-claims/claims-history.csv",
            "If new claim (not desk_rejected): call send_claim_under_review_email()",
            "get_dashboard_kpis(): reads all claim JSONs + index",
            "  KPIs computed: total claims, claims by date, claims by lossType,",
            "                 coverage breakdown, evidence confidence distribution,",
            "                 document type counts, desk-rejection rate",
            "clear_all_processed_claims(): wipe for dev/demo",
        ],
        "filters": "N/A",
        "policy_extraction": "N/A",
        "output": "data/processed-claims/*.json, claims-index.json, claims-history.csv; KPI JSON for frontend",
    },
    {
        "step": 10,
        "phase": "Claim Notifications (SMTP)",
        "service": "claim_notification",
        "file": "backend/claim_notification/service.py",
        "tech": "Python · smtplib · ssl · SMTP/SMTPS",
        "llm": "None",
        "summary": "Send automated email notifications to claimants at key workflow stages",
        "details": [
            "1. FNOL Receipt Acknowledgement (send_fnol_received_acknowledgement_email)",
            "   Triggered: after successful IMAP / SendGrid ingest",
            "   Toggle: FNOL_RECEIPT_ACK_EMAIL_ENABLED (default: on)",
            "   Content: 'We received your claim, reference #{claimId}'",
            "2. Claim Under Review (send_claim_under_review_email)",
            "   Triggered: after save_processed_claim() for non-desk-rejected claims",
            "   Toggle: CLAIM_UNDER_REVIEW_EMAIL_ENABLED (default: on)",
            "   Content: claim details, policy holder info, next steps",
            "3. Desk Rejection (send_desk_rejection_email)",
            "   Triggered: orchestrator detects expired/inactive/not-found policy",
            "   Toggle: DESK_REJECTION_EMAIL_ENABLED (default: on)",
            "   Content: rejection reason, policy status, appeal instructions",
            "SMTP config: SENDER_EMAIL, EMAIL_PASSWORD, SMTP_HOST, SMTP_PORT, SMTP_SECURE",
            "Reply threading: sets In-Reply-To / References headers matching original email",
        ],
        "filters": "N/A",
        "policy_extraction": "N/A",
        "output": "Outbound SMTP emails to claimant; threaded to original claim email",
    },
    {
        "step": 11,
        "phase": "FAQ Resolution",
        "service": "faq_resolution",
        "file": "backend/faq_resolution/service.py",
        "tech": "Python · OpenAI SDK · CSV · smtplib · regex",
        "llm": "gpt-4o-mini (FAQ vs CLAIM classification + FAQ answer index selection)",
        "summary": "Detect and automatically resolve FAQ / policy-info emails with SMTP reply",
        "details": [
            "process_faq_email(from, to, subject, body, message_id)",
            "1. Dedup: check faq-answered-ids.json + message_id → skip if already answered",
            "2. _is_faq_query(): regex layer (keyword patterns) + LLM ('FAQ or CLAIM? answer one word')",
            "3. If FAQ detected and SMTP configured:",
            "   a. _classify_customer_data_intent(): detect 'what is my policy limit / deductible / status'",
            "      → _answer_customer_data_query(): look up customer/policy JSON DB → personalised reply",
            "   b. Else: _find_faq_answer(): LLM receives full FAQ.csv context;",
            "      prompt asks to pick 'Q1', 'Q2', ... matching the question; extract answer text",
            "   c. Keyword fallback (no API key): overlap score between question and FAQ rows",
            "4. _send_faq_response_email(): SMTP threaded reply with In-Reply-To / References headers",
            "5. Persist message_id to faq-answered-ids.json to prevent duplicate replies",
            "FAQ CSV: data/FAQ.csv (question, answer pairs)",
        ],
        "filters": "_is_faq_query() gate; if classified as CLAIM → not handled here (goes to FNOL path)",
        "policy_extraction": "Reads policy data from local JSON DB for personalised answers",
        "output": "SMTP reply sent to sender; message ID stored to prevent duplicate replies",
    },
    {
        "step": 12,
        "phase": "FastAPI REST Layer",
        "service": "fastapi_server",
        "file": "backend/fastapi_server.py",
        "tech": "Python · FastAPI · Uvicorn · Pydantic v2 · CORS middleware",
        "llm": "None (delegates to services)",
        "summary": "REST API gateway exposing all backend services to the Next.js frontend",
        "details": [
            "Host: API_HOST (default 0.0.0.0), Port: API_PORT (default 8020)",
            "Hot reload: API_RELOAD env toggle",
            "CORS: CORS_ORIGINS (default * – allow all)",
            "Endpoints:",
            "  GET  /health → liveness probe",
            "  POST /api/process-claim → runs orchestrator pipeline",
            "  GET  /api/ingested-claims?full=bool&source=faq → list ingested claims",
            "  GET  /api/ingested-claims/{claim_id} → single ingested claim",
            "  POST /api/ingested-claims/clear → wipe ingested claims",
            "  POST /api/claims/clear → wipe all processed claims",
            "  GET  /api/claims → list processed claims",
            "  POST /api/claims → create/update processed claim",
            "  GET  /api/claims/{claim_id} → single processed claim",
            "  GET  /api/dashboard/kpis → KPI aggregation",
            "  POST /api/sync-inbox → trigger IMAP sync",
            "  GET  / → root health/info",
        ],
        "filters": "N/A",
        "policy_extraction": "N/A",
        "output": "JSON responses to frontend; all mutations persisted by downstream services",
    },
    {
        "step": 13,
        "phase": "Frontend – Next.js Application",
        "service": "frontend",
        "file": "frontend/app/page.tsx · components/*.tsx · lib/backend.ts",
        "tech": "Next.js 16 · React 18 · TypeScript · Tailwind CSS · Recharts · OpenAI npm · LangChain/LangGraph (deps present, main flow uses FastAPI)",
        "llm": "N/A (frontend calls FastAPI; LangGraph client-side orchestrator present but not used in main claim flow)",
        "summary": "Web UI for claim inbox, IMAP sync, claim processing, review, and analytics dashboard",
        "details": [
            "app/page.tsx: stage router – home → review → decision → dashboard; auth gate; FAQ page toggle",
            "HomePage.tsx: shows ingested claim inbox; IMAP sync via POST /api/sync-inbox; process claim via POST /api/process-claim",
            "ReviewPage.tsx: displays extracted fields + evidence list for adjuster review",
            "DecisionPage.tsx: shows decision pack – policy grounding, coverage confirmed, policy holder info",
            "DashboardPage.tsx: KPI charts via GET /api/dashboard/kpis; uses Recharts for visualisations",
            "FAQPage.tsx: FAQ-sourced inbox (source=faq query param)",
            "Header.tsx: navigation between pages",
            "ClaimSummaryBar.tsx: persistent claim summary across stages",
            "lib/backend.ts: runPython() helper for legacy webhook spawning",
            "lib/services/openai.ts: OpenAI client config (NEXT_PUBLIC_OPENAI_API_KEY)",
            "lib/agents/orchestrator.ts: LangGraph-style client-side orchestrator (not active in main flow)",
            "frontend/app/api/ingested-claims/route.ts: Next.js proxy route → FastAPI",
        ],
        "filters": "N/A",
        "policy_extraction": "N/A",
        "output": "Browser UI; all claim data read/written via FastAPI",
    },
]

ENV_VARS = [
    ("OPENAI_API_KEY", "OpenAI", "Required for LLM calls (gpt-4o / gpt-4o-mini)"),
    ("OPENAI_MODEL", "OpenAI", "Default: gpt-4o (extraction); gpt-4o-mini (FNOL gate, FAQ)"),
    ("IMAP_HOST", "IMAP", "Default: imap.gmail.com"),
    ("IMAP_PORT", "IMAP", "Default: 993"),
    ("SENDER_EMAIL / IMAP_USER", "IMAP", "Inbox email address"),
    ("EMAIL_PASSWORD / IMAP_PASSWORD", "IMAP", "App password (spaces stripped)"),
    ("IMAP_MAILBOX", "IMAP", "Default: INBOX"),
    ("IMAP_SYNC_INCLUDE_READ", "IMAP", "Default: true – fetch ALL, not just UNSEEN"),
    ("IMAP_SYNC_MAX_EMAILS", "IMAP", "Default: 0 (unlimited)"),
    ("IMAP_GMAIL_TRY_ALL_MAIL", "IMAP", "Default: true – try [Gmail]/All Mail first"),
    ("IMAP_SSL_VERIFY", "IMAP", "Default: false – skip cert verification"),
    ("FNOL_FILTER_ENABLED", "FNOL Gate", "Default: true – use LLM FNOL classification"),
    ("FNOL_VERBOSE_SKIP_LOGS", "FNOL Gate", "Verbose logging for skipped emails"),
    ("FNOL_RECEIPT_ACK_EMAIL_ENABLED", "Notifications", "Default: on – send FNOL ack"),
    ("CLAIM_UNDER_REVIEW_EMAIL_ENABLED", "Notifications", "Default: on – send under-review email"),
    ("DESK_REJECTION_EMAIL_ENABLED", "Notifications", "Default: on – send desk-rejection email"),
    ("SMTP_HOST", "SMTP", "Outbound SMTP host"),
    ("SMTP_PORT", "SMTP", "Outbound SMTP port"),
    ("SMTP_SECURE", "SMTP", "TLS/SSL flag"),
    ("API_HOST", "FastAPI", "Default: 0.0.0.0"),
    ("API_PORT", "FastAPI", "Default: 8020"),
    ("API_RELOAD", "FastAPI", "Uvicorn hot reload toggle"),
    ("CORS_ORIGINS", "FastAPI", "Default: * (allow all origins)"),
    ("LOCAL_DATA_DIR", "Policy DB", "Default: database/local_data"),
    ("POLICY_GROUNDING_MAPPING_FILE", "Policy DB", "Default: database/policy_grounding_mapping.json"),
    ("FAQ_CSV_FILE", "FAQ", "Default: data/FAQ.csv"),
    ("DATA_DIR / INGESTED_DIR / PROCESSED_CLAIMS_DIR", "Paths", "Data directory overrides"),
    ("ENV_FILE", "Config", "Path to .env file"),
    ("NEXT_PUBLIC_API_URL", "Frontend", "FastAPI base URL (default: http://localhost:8020)"),
    ("NEXT_PUBLIC_OPENAI_API_KEY", "Frontend", "OpenAI key for client-side SDK (legacy flows)"),
]

ENDPOINTS = [
    ("GET",  "/health",                          "Liveness probe",                                                      "fastapi_server"),
    ("GET",  "/",                                "Root info endpoint",                                                  "fastapi_server"),
    ("POST", "/api/process-claim",               "Run full orchestration pipeline on ingested claim",                   "process_claim/orchestrator"),
    ("GET",  "/api/ingested-claims",             "List all ingested claims; query: full=bool, source=faq",              "ingested_claims/service"),
    ("GET",  "/api/ingested-claims/{claim_id}",  "Get single ingested claim by ID",                                    "ingested_claims/service"),
    ("POST", "/api/ingested-claims/clear",       "Wipe all ingested claims (dev/demo reset)",                          "ingested_claims/service"),
    ("POST", "/api/claims/clear",                "Wipe all processed claims (dev/demo reset)",                         "dashboard/service"),
    ("GET",  "/api/claims",                      "List all processed claims",                                           "dashboard/service"),
    ("POST", "/api/claims",                      "Create / update processed claim record",                             "dashboard/service"),
    ("GET",  "/api/claims/{claim_id}",           "Get single processed claim by ID",                                   "dashboard/service"),
    ("GET",  "/api/dashboard/kpis",              "Return KPI aggregations for dashboard charts",                       "dashboard/service"),
    ("POST", "/api/sync-inbox",                  "Trigger IMAP inbox sync (runs email_ingestion pipeline)",            "email_ingestion/service"),
]

DATA_STORES = [
    ("data/ingested-claims.json", "Ingested claims", "All raw ingested emails (FNOL + FAQ)"),
    ("data/ingested-attachments/{claimId}/", "Attachment files", "Binary/text attachments per claim"),
    ("data/faq-answered-ids.json", "FAQ dedup list", "Message IDs of already-answered FAQ emails"),
    ("data/processed-claims/{claimId}.json", "Processed claim", "Full decision pack per claim"),
    ("data/processed-claims/claims-index.json", "Claims index", "Lightweight index of all processed claims"),
    ("data/processed-claims/claims-history.csv", "Claims CSV", "Append-only CSV for audit/analytics"),
    ("database/local_data/customers.json", "Customer DB", "Customer records (name, email, address)"),
    ("database/local_data/policies.json", "Policy DB", "Policy records (type, status, limits, expiry)"),
    ("database/local_data/policy_details.json", "Policy details DB", "Detailed clause and coverage info"),
    ("database/policy_grounding_mapping.json", "Coverage mapping", "lossType × policy_type → clause IDs"),
    ("data/FAQ.csv", "FAQ dataset", "Question-answer pairs for FAQ resolution"),
]

LLM_SUMMARY = [
    ("gpt-4o", "Extraction", "email_extraction", "Extract structured fields from FNOL email body", "temp=0.1, max_tokens=1500"),
    ("gpt-4o", "Extraction", "document_extraction", "Classify attachment type and extract key fields", "temp=0.1, max_tokens=1500"),
    ("gpt-4o", "Extraction", "vision_analysis", "Analyse damage photos, multimodal image_url data URL", "max_tokens=2500"),
    ("gpt-4o-mini", "Email Ingestion", "fnol_classifier", "Binary yes/no: is this email a P&C FNOL? Answer only yes or no.", "max_tokens=10, temp=0"),
    ("gpt-4o-mini", "FAQ Resolution", "faq_vs_claim", "One-word classification: FAQ or CLAIM?", "default settings"),
    ("gpt-4o-mini", "FAQ Resolution", "faq_answer_selector", "Pick matching Q-index (Q1, Q2, …) from FAQ.csv context; extract answer text", "default settings"),
]


# ─────────────────────────────────────────────────────────
# XLSX  GENERATION
# ─────────────────────────────────────────────────────────

def hex_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color="FFFFFF"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="1A1A2E"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def apply_header(ws, row, values, fill_hex, font_color="FFFFFF", font_size=11):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = hex_fill(fill_hex)
        c.font = header_font(size=font_size, color=font_color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()

def apply_row(ws, row, values, fill_hex=None, bold=False, wrap=True):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        if fill_hex:
            c.fill = hex_fill(fill_hex)
        c.font = body_font(bold=bold)
        c.alignment = Alignment(vertical="top", wrap_text=wrap)
        c.border = thin_border()

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


def build_xlsx(out_path):
    wb = Workbook()

    # ── Sheet 1: Overview ─────────────────────────────────
    ws1 = wb.active
    ws1.title = "Overview"
    ws1.sheet_view.showGridLines = False
    ws1.row_dimensions[1].height = 40
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = "Autonomous Claims Orchestrator – End-to-End Workflow Overview"
    title_cell.fill = hex_fill("1A237E")
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Step", "Phase", "Service / Module", "Primary File(s)", "Tech Stack", "LLM Used", "Summary"]
    apply_header(ws1, 2, headers, "283593", font_size=10)
    freeze(ws1, "A3")

    PHASE_COLORS = {
        "Email Ingestion": "E8F5E9",
        "SendGrid Inbound (Alternate Ingestion)": "F3E5F5",
        "Ingested Claims Storage": "FFF8E1",
        "Claim Extraction": "E3F2FD",
        "Policy Grounding (Local DB)": "FCE4EC",
        "Policy Grounding (ISO Clause Fallback)": "FBE9E7",
        "Decision Pack Assembly": "E8EAF6",
        "Process Claim Orchestration": "F1F8E9",
        "Dashboard & Processed Claims Storage": "E0F7FA",
        "Claim Notifications (SMTP)": "FFF3E0",
        "FAQ Resolution": "F9FBE7",
        "FastAPI REST Layer": "EDE7F6",
        "Frontend – Next.js Application": "E8EAF6",
    }

    for i, s in enumerate(WORKFLOW_STEPS):
        r = i + 3
        ws1.row_dimensions[r].height = 50
        fill = PHASE_COLORS.get(s["phase"], "FAFAFA")
        apply_row(ws1, r, [
            s["step"], s["phase"], s["service"], s["file"], s["tech"], s["llm"], s["summary"]
        ], fill_hex=fill)

    set_col_widths(ws1, [6, 26, 22, 52, 50, 44, 50])

    # ── Sheet 2: Detailed Steps ────────────────────────────
    ws2 = wb.create_sheet("Detailed Steps")
    ws2.sheet_view.showGridLines = False
    ws2.row_dimensions[1].height = 36
    ws2.merge_cells("A1:D1")
    t = ws2["A1"]
    t.value = "Step-by-Step Details – All Microservices"
    t.fill = hex_fill("0D47A1")
    t.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")

    apply_header(ws2, 2, ["Step / Phase", "LLM Model", "Details (Bullet Points)", "Output"], "1565C0", font_size=10)
    freeze(ws2, "A3")

    row = 3
    for s in WORKFLOW_STEPS:
        details_text = "\n".join(f"• {d}" for d in s["details"])
        ws2.row_dimensions[row].height = max(60, len(s["details"]) * 15)
        fill = PHASE_COLORS.get(s["phase"], "FAFAFA")
        apply_row(ws2, row, [
            f"[{s['step']}] {s['phase']}\n{s['service']}",
            s["llm"],
            details_text,
            s["output"],
        ], fill_hex=fill)
        row += 1

    set_col_widths(ws2, [30, 40, 90, 55])

    # ── Sheet 3: Email Ingestion Deep-Dive ───────────────
    ws3 = wb.create_sheet("Email Ingestion Deep-Dive")
    ws3.sheet_view.showGridLines = False
    ws3.merge_cells("A1:C1")
    t = ws3["A1"]
    t.value = "Email Ingestion – IMAP Filters, Policy Extraction, MIME Parsing"
    t.fill = hex_fill("1B5E20")
    t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    apply_header(ws3, 2, ["Category", "Item", "Detail"], "2E7D32", font_size=10)
    freeze(ws3, "A3")

    ingestion_rows = [
        ("IMAP Config", "Host", "imap.gmail.com (env: IMAP_HOST)"),
        ("IMAP Config", "Port", "993 (env: IMAP_PORT)"),
        ("IMAP Config", "Auth", "SENDER_EMAIL / IMAP_USER + EMAIL_PASSWORD / IMAP_PASSWORD (spaces stripped)"),
        ("IMAP Config", "Mailbox", "INBOX (env: IMAP_MAILBOX); Gmail: tries [Gmail]/All Mail first"),
        ("IMAP Config", "SSL", "IMAP4_SSL; cert verify OFF by default (IMAP_SSL_VERIFY=false)"),
        ("IMAP Config", "Fetch scope", "SEARCH ALL (include read) or UNSEEN; controlled by IMAP_SYNC_INCLUDE_READ"),
        ("IMAP Config", "Max emails", "IMAP_SYNC_MAX_EMAILS=0 → unlimited"),
        ("Filter 1", "Duplicate check", "Message-ID match → skip; fallback: subject|from|date fingerprint hash"),
        ("Filter 2", "Thread reply", "If reply to existing claim → merge into thread, skip FNOL re-classification"),
        ("Filter 3", "Consumer complaint", "_should_reject_consumer_complaint_correspondence(): OEM/retail patterns, short replies on quoted threads → skip FNOL"),
        ("Filter 4", "Procedural FAQ", "_is_clearly_procedural_faq_question(): 'how do I file', expiry questions → route to FAQ service"),
        ("Filter 5", "Keyword gate", "_has_relevant_keywords() + _has_strong_keywords() on _primary_message_text() (latest non-quoted segment)"),
        ("Filter 6", "LLM FNOL gate", "gpt-4o-mini; system prompt: 'Is this a P&C FNOL? Answer only yes or no'; temp=0, max_tokens=10; bypassed if FNOL_FILTER_ENABLED=false"),
        ("Policy # Extraction", "Pattern group 1 – keyword-anchored", r"poli?cy\s*(?:number|no\.?|#|id|ref(?:erence)?)?\s*[:#\-=\s]\s*([A-Z0-9]{4,}(?:[-/][A-Z0-9]+)*)"),
        ("Policy # Extraction", "Pattern group 1 – claim number", r"claim\s*(?:number|no\.?|#|id|ref)?\s*[:#\s]\s*([A-Z0-9]{4,}(?:[-/][A-Z0-9]+)*)"),
        ("Policy # Extraction", "Pattern group 1 – hash prefix", r"#([A-Z]{1,4}\d{5,})"),
        ("Policy # Extraction", "Pattern group 2 – structural", r"\b([A-Z]{2}\d{6,})\b  |  \b([A-Z]{3}\d{6,})\b  |  \b([A-Z]{2,4}-\d{5,})\b"),
        ("Policy # Extraction", "Fallback", "If no regex match → sender email / Message-ID display used as policyNumber placeholder"),
        ("MIME Parsing", "Body extraction", "_extract_body_text(): prefers first text/plain part; fallback text/html → _strip_html()"),
        ("MIME Parsing", "Headers", "_decode_header_value() (MIME-encoded words), _format_address()"),
        ("MIME Parsing", "Attachments", "Collected as {'filename', 'content_type', 'data': bytes}; stored to data/ingested-attachments/{claimId}/"),
        ("MIME Parsing", "SendGrid path", "extract_plain_body_from_rfc822(): parses raw RFC822 email field when text/html fields empty"),
    ]

    for r_idx, (cat, item, detail) in enumerate(ingestion_rows, 3):
        fill = "E8F5E9" if r_idx % 2 == 0 else "F9FFF9"
        apply_row(ws3, r_idx, [cat, item, detail], fill_hex=fill)

    set_col_widths(ws3, [26, 36, 85])

    # ── Sheet 4: Extraction Deep-Dive ─────────────────────
    ws4 = wb.create_sheet("Extraction Deep-Dive")
    ws4.sheet_view.showGridLines = False
    ws4.merge_cells("A1:C1")
    t = ws4["A1"]
    t.value = "Claim Extraction – LLM Prompts, Fields, Document & Vision Analysis"
    t.fill = hex_fill("0D47A1")
    t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    apply_header(ws4, 2, ["Extraction Type", "Parameter / Field", "Value / Detail"], "1565C0", font_size=10)
    freeze(ws4, "A3")

    extraction_rows = [
        ("Email Extraction", "Model", "gpt-4o (env: OPENAI_MODEL)"),
        ("Email Extraction", "Temperature", "0.1"),
        ("Email Extraction", "Max tokens", "1500"),
        ("Email Extraction", "System prompt", "You are an expert insurance claims processor. Extract the following information from the FNOL email. Return ONLY valid JSON."),
        ("Email Extraction", "Field: policyNumber", "Policy number string from email body/subject"),
        ("Email Extraction", "Field: claimantName", "Full name of the person filing the claim"),
        ("Email Extraction", "Field: contactEmail", "Contact email address of claimant"),
        ("Email Extraction", "Field: contactPhone", "Contact phone number"),
        ("Email Extraction", "Field: lossDate", "Date of loss/incident (ISO format preferred)"),
        ("Email Extraction", "Field: lossType", "Enum: fire | water | auto | theft | liability | medical | other"),
        ("Email Extraction", "Field: lossLocation", "Address or description of incident location"),
        ("Email Extraction", "Field: description", "Full narrative description of the loss event"),
        ("Email Extraction", "Field: vehicleInfo", "Vehicle make/model/year/VIN if auto claim"),
        ("Email Extraction", "Field: propertyAddress", "Property address if home/commercial claim"),
        ("Email Extraction", "Field: estimatedAmount", "Claimant's estimated loss amount (numeric)"),
        ("Email Extraction", "Field: _confidence", "Per-field confidence map (0.0–1.0) indicating extraction certainty"),
        ("Document Extraction", "Model", "gpt-4o (same env var)"),
        ("Document Extraction", "Temperature", "0.1"),
        ("Document Extraction", "Max tokens", "1500"),
        ("Document Extraction", "Document types", "PoliceReport | RepairEstimate | Invoice | MedicalRecord | IncidentReport | DamagePhoto | Other"),
        ("Document Extraction", "Logic", "Classify attachment type; use _extract_schemas() for type-specific field hints; extract keyFields"),
        ("Vision Analysis", "Model", "gpt-4o (multimodal)"),
        ("Vision Analysis", "Max tokens", "2500"),
        ("Vision Analysis", "Input", "Image bytes base64-encoded → image_url data URL in message content"),
        ("Vision Analysis", "Output fields", "damage_severity, repair_cost_estimate, injury_indicators, detailed_summary, additional_observations"),
        ("Evidence Assembly", "Logic", "For each scalar extractedField value + confidence score → build evidence[] entry"),
        ("Evidence Assembly", "Evidence item shape", "{ field, value, confidence, source: 'email'|'document'|'vision' }"),
    ]

    for r_idx, (etype, param, val) in enumerate(extraction_rows, 3):
        fill = "E3F2FD" if r_idx % 2 == 0 else "F0F8FF"
        apply_row(ws4, r_idx, [etype, param, val], fill_hex=fill)

    set_col_widths(ws4, [26, 32, 90])

    # ── Sheet 5: LLM Usage ────────────────────────────────
    ws5 = wb.create_sheet("LLM Usage")
    ws5.sheet_view.showGridLines = False
    ws5.merge_cells("A1:F1")
    t = ws5["A1"]
    t.value = "All LLM Calls – Models, Services, Prompts, Parameters"
    t.fill = hex_fill("4A148C")
    t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 32

    apply_header(ws5, 2, ["Model", "Service", "Call Name", "Purpose / Prompt", "Parameters", "Bypass Condition"], "6A1B9A", font_size=10)
    freeze(ws5, "A3")

    llm_rows = [
        ("gpt-4o", "extraction", "extract_from_email", "Extract structured JSON fields from FNOL email body (policyNumber, claimantName, lossType, etc.)", "temp=0.1, max_tokens=1500", "None – always runs"),
        ("gpt-4o", "extraction", "extract_from_document", "Classify attachment document type and extract type-specific key fields", "temp=0.1, max_tokens=1500", "None – runs per text attachment"),
        ("gpt-4o (vision)", "extraction", "analyze_image_with_vision", "Multimodal: analyse damage photo for severity, repair cost, injury indicators", "max_tokens=2500", "None – runs per image attachment"),
        ("gpt-4o-mini", "email_ingestion", "should_ingest_incoming_email (FNOL gate)", "System: 'Is this email a P&C First Notice of Loss? Answer only yes or no.' User: full email body", "temp=0, max_tokens=10", "FNOL_FILTER_ENABLED=false OR strong keyword match with no API key"),
        ("gpt-4o-mini", "faq_resolution", "_is_faq_query (FAQ vs CLAIM)", "One-word classification: is this email a FAQ question or a CLAIM filing?", "default settings", "Regex pre-filter; if clear regex match, LLM not needed"),
        ("gpt-4o-mini", "faq_resolution", "_find_faq_answer (FAQ answer selector)", "Receives full FAQ.csv text; prompted to return matching Q-index (Q1, Q2, …); answer text extracted from CSV row", "default settings", "No API key → keyword overlap fallback"),
    ]

    for r_idx, row_data in enumerate(llm_rows, 3):
        fill = "F3E5F5" if r_idx % 2 == 0 else "FAFAFA"
        apply_row(ws5, r_idx, list(row_data), fill_hex=fill)

    set_col_widths(ws5, [18, 20, 36, 72, 32, 40])

    # ── Sheet 6: API Endpoints ────────────────────────────
    ws6 = wb.create_sheet("API Endpoints")
    ws6.sheet_view.showGridLines = False
    ws6.merge_cells("A1:D1")
    t = ws6["A1"]
    t.value = "FastAPI Endpoints – Complete Reference"
    t.fill = hex_fill("BF360C")
    t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 32

    apply_header(ws6, 2, ["Method", "Path", "Purpose", "Backend Service Called"], "D84315", font_size=10)
    freeze(ws6, "A3")

    for r_idx, ep in enumerate(ENDPOINTS, 3):
        method_colors = {"GET": "E8F5E9", "POST": "FFF3E0", "DELETE": "FCE4EC"}
        fill = method_colors.get(ep[0], "FAFAFA")
        apply_row(ws6, r_idx, list(ep), fill_hex=fill)

    set_col_widths(ws6, [10, 42, 55, 32])

    # ── Sheet 7: Data Stores ──────────────────────────────
    ws7 = wb.create_sheet("Data Stores")
    ws7.sheet_view.showGridLines = False
    ws7.merge_cells("A1:C1")
    t = ws7["A1"]
    t.value = "Data Storage – All Files, Formats, and Purposes"
    t.fill = hex_fill("004D40")
    t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws7.row_dimensions[1].height = 32

    apply_header(ws7, 2, ["File / Directory", "Format / Role", "Description"], "00695C", font_size=10)
    freeze(ws7, "A3")

    for r_idx, (path, role, desc) in enumerate(DATA_STORES, 3):
        fill = "E0F2F1" if r_idx % 2 == 0 else "F1FFFE"
        apply_row(ws7, r_idx, [path, role, desc], fill_hex=fill)

    set_col_widths(ws7, [48, 28, 65])

    # ── Sheet 8: Environment Variables ───────────────────
    ws8 = wb.create_sheet("Environment Variables")
    ws8.sheet_view.showGridLines = False
    ws8.merge_cells("A1:C1")
    t = ws8["A1"]
    t.value = "Environment Variables – Complete Reference"
    t.fill = hex_fill("1A237E")
    t.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws8.row_dimensions[1].height = 32

    apply_header(ws8, 2, ["Variable(s)", "Category", "Description / Default"], "283593", font_size=10)
    freeze(ws8, "A3")

    for r_idx, (var, cat, desc) in enumerate(ENV_VARS, 3):
        fill = "E8EAF6" if r_idx % 2 == 0 else "F5F5FF"
        apply_row(ws8, r_idx, [var, cat, desc], fill_hex=fill)

    set_col_widths(ws8, [42, 20, 70])

    wb.save(out_path)
    print(f"XLSX saved → {out_path}")


# ─────────────────────────────────────────────────────────
# PDF  GENERATION
# ─────────────────────────────────────────────────────────

from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable, KeepTogether
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY

# Colour palette
C_NAVY     = colors.HexColor("#1A237E")
C_BLUE     = colors.HexColor("#1565C0")
C_LBLUE    = colors.HexColor("#E3F2FD")
C_GREEN    = colors.HexColor("#1B5E20")
C_LGREEN   = colors.HexColor("#E8F5E9")
C_PURPLE   = colors.HexColor("#4A148C")
C_LPURPLE  = colors.HexColor("#F3E5F5")
C_ORANGE   = colors.HexColor("#BF360C")
C_LORANGE  = colors.HexColor("#FFF3E0")
C_TEAL     = colors.HexColor("#004D40")
C_LTEAL    = colors.HexColor("#E0F2F1")
C_WHITE    = colors.white
C_LGRAY    = colors.HexColor("#F5F5F5")
C_GRAY     = colors.HexColor("#EEEEEE")
C_DGRAY    = colors.HexColor("#424242")
C_PHASE    = {
    "Email Ingestion":                             colors.HexColor("#E8F5E9"),
    "SendGrid Inbound (Alternate Ingestion)":      colors.HexColor("#F3E5F5"),
    "Ingested Claims Storage":                     colors.HexColor("#FFF8E1"),
    "Claim Extraction":                            colors.HexColor("#E3F2FD"),
    "Policy Grounding (Local DB)":                 colors.HexColor("#FCE4EC"),
    "Policy Grounding (ISO Clause Fallback)":      colors.HexColor("#FBE9E7"),
    "Decision Pack Assembly":                      colors.HexColor("#E8EAF6"),
    "Process Claim Orchestration":                 colors.HexColor("#F1F8E9"),
    "Dashboard & Processed Claims Storage":        colors.HexColor("#E0F7FA"),
    "Claim Notifications (SMTP)":                  colors.HexColor("#FFF3E0"),
    "FAQ Resolution":                              colors.HexColor("#F9FBE7"),
    "FastAPI REST Layer":                          colors.HexColor("#EDE7F6"),
    "Frontend – Next.js Application":             colors.HexColor("#E8EAF6"),
}


def build_styles():
    base = getSampleStyleSheet()
    styles = {}
    styles["title"] = ParagraphStyle(
        "title", parent=base["Title"],
        fontSize=22, textColor=C_WHITE, alignment=TA_CENTER,
        spaceAfter=6, fontName="Helvetica-Bold",
    )
    styles["subtitle"] = ParagraphStyle(
        "subtitle", parent=base["Normal"],
        fontSize=12, textColor=C_WHITE, alignment=TA_CENTER,
        spaceAfter=4, fontName="Helvetica",
    )
    styles["h1"] = ParagraphStyle(
        "h1", parent=base["Heading1"],
        fontSize=14, textColor=C_WHITE, fontName="Helvetica-Bold",
        spaceAfter=4, spaceBefore=8,
    )
    styles["h2"] = ParagraphStyle(
        "h2", parent=base["Heading2"],
        fontSize=12, textColor=C_NAVY, fontName="Helvetica-Bold",
        spaceBefore=8, spaceAfter=4,
    )
    styles["h3"] = ParagraphStyle(
        "h3", parent=base["Heading3"],
        fontSize=10, textColor=C_BLUE, fontName="Helvetica-Bold",
        spaceBefore=4, spaceAfter=2,
    )
    styles["body"] = ParagraphStyle(
        "body", parent=base["Normal"],
        fontSize=9, textColor=C_DGRAY, fontName="Helvetica",
        leading=13, spaceAfter=2,
    )
    styles["bullet"] = ParagraphStyle(
        "bullet", parent=base["Normal"],
        fontSize=8.5, textColor=C_DGRAY, fontName="Helvetica",
        leading=12, leftIndent=12, firstLineIndent=-8, spaceAfter=1,
    )
    styles["code"] = ParagraphStyle(
        "code", parent=base["Code"],
        fontSize=7.5, fontName="Courier",
        textColor=colors.HexColor("#333333"),
        backColor=colors.HexColor("#F5F5F5"),
        leftIndent=8, rightIndent=8, leading=11,
    )
    styles["cell"] = ParagraphStyle(
        "cell", parent=base["Normal"],
        fontSize=8, fontName="Helvetica", leading=10,
    )
    styles["cell_hdr"] = ParagraphStyle(
        "cell_hdr", parent=base["Normal"],
        fontSize=9, fontName="Helvetica-Bold",
        textColor=C_WHITE, leading=11,
    )
    return styles


def section_banner(text, bg_color, styles):
    data = [[Paragraph(text, styles["h1"])]]
    t = Table(data, colWidths=["100%"])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), bg_color),
        ("TOPPADDING",    (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("LEFTPADDING",   (0, 0), (-1, -1), 12),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 12),
    ]))
    return t


def kv_table(rows, styles, col_w=(55*mm, None)):
    """2-column key-value table."""
    page_w = A4[0] - 40*mm
    w2 = page_w - col_w[0]
    data = []
    for k, v in rows:
        data.append([
            Paragraph(f"<b>{k}</b>", styles["cell"]),
            Paragraph(v, styles["cell"]),
        ])
    t = Table(data, colWidths=[col_w[0], w2])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#ECEFF1")),
        ("BACKGROUND", (1, 0), (1, -1), C_WHITE),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#B0BEC5")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 6),
    ]))
    return t


def build_pdf(out_path):
    doc = SimpleDocTemplate(
        out_path,
        pagesize=A4,
        rightMargin=20*mm, leftMargin=20*mm,
        topMargin=18*mm, bottomMargin=18*mm,
        title="Autonomous Claims Orchestrator – End-to-End Workflow",
        author="Claims Engineering",
    )
    styles = build_styles()
    story = []
    page_w = A4[0] - 40*mm

    # ── Cover ─────────────────────────────────────────────
    cover_data = [[
        Paragraph("Autonomous Claims Orchestrator", styles["title"]),
        Paragraph("End-to-End Workflow Documentation", styles["subtitle"]),
        Paragraph("Tech Stacks · Microservices · LLM Models · Data Flows", styles["subtitle"]),
    ]]
    cover_table = Table([[d] for d in cover_data[0]], colWidths=[page_w])
    cover_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), C_NAVY),
        ("TOPPADDING",    (0, 0), (-1, -1), 18),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 18),
        ("LEFTPADDING",   (0, 0), (-1, -1), 16),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 16),
    ]))
    story.append(cover_table)
    story.append(Spacer(1, 10*mm))

    # Quick facts
    qf_data = [
        ["Backend", "Python 3 · FastAPI · Uvicorn · Pydantic v2"],
        ["Frontend", "Next.js 16 · React 18 · TypeScript · Tailwind · Recharts"],
        ["LLM (extraction)", "gpt-4o (email, document, vision)"],
        ["LLM (classification)", "gpt-4o-mini (FNOL gate, FAQ detection, FAQ answer)"],
        ["Email ingest", "IMAP4_SSL (imap.gmail.com:993) + SendGrid Inbound Parse webhook"],
        ["Storage", "JSON file store + CSV (no SQL at runtime)"],
        ["Notifications", "smtplib SMTP (FNOL ack, under-review, desk-rejection)"],
        ["Microservices", "email_ingestion · ingested_claims · extraction · decision · faq_resolution · process_claim · dashboard · claim_notification"],
    ]
    qf_table = Table(qf_data, colWidths=[50*mm, page_w - 50*mm])
    qf_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), C_NAVY),
        ("BACKGROUND", (1, 0), (1, -1), C_LBLUE),
        ("TEXTCOLOR",  (0, 0), (0, -1), C_WHITE),
        ("FONTNAME",   (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME",   (1, 0), (1, -1), "Helvetica"),
        ("FONTSIZE",   (0, 0), (-1, -1), 9),
        ("GRID",       (0, 0), (-1, -1), 0.3, C_BLUE),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING",   (0, 0), (-1, -1), 8),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(qf_table)
    story.append(PageBreak())

    # ── Section 1: Workflow steps (overview table) ────────
    story.append(section_banner("1. End-to-End Workflow Overview", C_NAVY, styles))
    story.append(Spacer(1, 4*mm))

    ov_headers = [
        Paragraph("<b>Step</b>", styles["cell_hdr"]),
        Paragraph("<b>Phase</b>", styles["cell_hdr"]),
        Paragraph("<b>Service</b>", styles["cell_hdr"]),
        Paragraph("<b>Tech Stack</b>", styles["cell_hdr"]),
        Paragraph("<b>LLM</b>", styles["cell_hdr"]),
    ]
    ov_data = [ov_headers]
    for s in WORKFLOW_STEPS:
        row = [
            Paragraph(str(s["step"]), styles["cell"]),
            Paragraph(f"<b>{s['phase']}</b>", styles["cell"]),
            Paragraph(s["service"], styles["cell"]),
            Paragraph(s["tech"], styles["cell"]),
            Paragraph(s["llm"], styles["cell"]),
        ]
        ov_data.append(row)

    ov_table = Table(ov_data, colWidths=[10*mm, 48*mm, 38*mm, 50*mm, 44*mm])
    ov_style = [
        ("BACKGROUND", (0, 0), (-1, 0), C_NAVY),
        ("TEXTCOLOR",  (0, 0), (-1, 0), C_WHITE),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.HexColor("#B0BEC5")),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("VALIGN",     (0, 0), (-1, -1), "TOP"),
    ]
    for i, s in enumerate(WORKFLOW_STEPS, 1):
        phase_c = C_PHASE.get(s["phase"], C_LGRAY)
        ov_style.append(("BACKGROUND", (0, i), (-1, i), phase_c))
    ov_table.setStyle(TableStyle(ov_style))
    story.append(ov_table)
    story.append(PageBreak())

    # ── Section 2: Detailed per-step narrative ───────────
    story.append(section_banner("2. Detailed Step-by-Step Workflow", C_NAVY, styles))
    story.append(Spacer(1, 4*mm))

    for s in WORKFLOW_STEPS:
        phase_c = C_PHASE.get(s["phase"], C_LGRAY)
        block = []

        # Step header
        hdr_data = [[Paragraph(
            f"Step {s['step']} &nbsp;|&nbsp; {s['phase']}",
            ParagraphStyle("step_hdr", parent=styles["h2"],
                           textColor=C_WHITE, fontSize=11, fontName="Helvetica-Bold")
        )]]
        hdr_t = Table(hdr_data, colWidths=[page_w])
        hdr_t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), C_BLUE),
            ("TOPPADDING",    (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ]))
        block.append(hdr_t)
        block.append(Spacer(1, 2*mm))

        kv = [
            ("Service / Module", s["service"]),
            ("File(s)", s["file"]),
            ("Tech Stack", s["tech"]),
            ("LLM Model", s["llm"]),
            ("Output", s["output"]),
        ]
        block.append(kv_table(kv, styles))
        block.append(Spacer(1, 2*mm))

        block.append(Paragraph("What Happens (Step by Step):", styles["h3"]))
        for d in s["details"]:
            block.append(Paragraph(f"• {d}", styles["bullet"]))
        block.append(Spacer(1, 1*mm))

        if s.get("filters") and s["filters"] != "N/A":
            block.append(Paragraph("Filters Applied:", styles["h3"]))
            if isinstance(s["filters"], list):
                for f in s["filters"]:
                    block.append(Paragraph(f"• {f}", styles["bullet"]))
            else:
                block.append(Paragraph(s["filters"], styles["body"]))
            block.append(Spacer(1, 1*mm))

        if s.get("policy_extraction") and s["policy_extraction"] not in ("N/A", ""):
            block.append(Paragraph("Policy Number Handling:", styles["h3"]))
            block.append(Paragraph(s["policy_extraction"], styles["body"]))
            block.append(Spacer(1, 1*mm))

        block.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CFD8DC")))
        block.append(Spacer(1, 3*mm))

        story.append(KeepTogether(block[:4]))
        for item in block[4:]:
            story.append(item)

    story.append(PageBreak())

    # ── Section 3: LLM Usage ─────────────────────────────
    story.append(section_banner("3. LLM Models – Detailed Usage", C_PURPLE, styles))
    story.append(Spacer(1, 4*mm))

    llm_h = [Paragraph(h, styles["cell_hdr"]) for h in
              ["Model", "Service", "Call", "Purpose", "Parameters", "Bypass"]]
    llm_data = [llm_h]
    for row in LLM_SUMMARY:
        llm_data.append([Paragraph(v, styles["cell"]) for v in row])

    llm_t = Table(llm_data, colWidths=[20*mm, 22*mm, 28*mm, 60*mm, 24*mm, 36*mm])
    llm_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), C_PURPLE),
        ("GRID",       (0,0), (-1,-1), 0.3, colors.HexColor("#CE93D8")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [C_LPURPLE, C_WHITE]),
        ("TOPPADDING",    (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("VALIGN",     (0,0), (-1,-1), "TOP"),
    ]))
    story.append(llm_t)
    story.append(PageBreak())

    # ── Section 4: Email Ingestion deep-dive ─────────────
    story.append(section_banner("4. Email Ingestion – Deep Dive", C_GREEN, styles))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("4.1 IMAP Configuration", styles["h2"]))
    imap_kv = [
        ("Host (default)", "imap.gmail.com  [env: IMAP_HOST]"),
        ("Port (default)", "993  [env: IMAP_PORT]"),
        ("Protocol", "IMAP4_SSL; SSL certificate verify OFF by default (IMAP_SSL_VERIFY=false)"),
        ("Auth", "SENDER_EMAIL / IMAP_USER  +  EMAIL_PASSWORD / IMAP_PASSWORD (spaces automatically stripped)"),
        ("Mailbox", "INBOX  [env: IMAP_MAILBOX];  Gmail: tries [Gmail]/All Mail → [Google Mail]/All Mail → INBOX"),
        ("Fetch scope", "SEARCH ALL (IMAP_SYNC_INCLUDE_READ=true, default) or SEARCH UNSEEN"),
        ("Max emails", "IMAP_SYNC_MAX_EMAILS=0 → unlimited"),
    ]
    story.append(kv_table(imap_kv, styles))
    story.append(Spacer(1, 3*mm))

    story.append(Paragraph("4.2 Filters Applied (Application-Side)", styles["h2"]))
    filters = [
        ("1. Duplicate", "Message-ID exact match → skip. Fallback fingerprint: SHA of subject|from|date when no Message-ID."),
        ("2. Thread reply", "If in-reply-to matches existing claim ID → merge email into thread; skip FNOL re-classification entirely."),
        ("3. Consumer complaint", "_should_reject_consumer_complaint_correspondence(): detects OEM / retail complaint patterns, short replies on quoted threads → classified as non-FNOL, dropped."),
        ("4. Procedural FAQ", "_is_clearly_procedural_faq_question(): regex detects 'how do I file', 'when does my policy expire', etc. → routed to FAQ service, not FNOL."),
        ("5. Keyword gate", "_has_relevant_keywords() + _has_strong_keywords() run on _primary_message_text() (latest non-quoted segment). Strong match can bypass LLM gate."),
        ("6. LLM FNOL gate", "gpt-4o-mini, temp=0, max_tokens=10. System: 'You classify P&C insurance emails. Is this a First Notice of Loss? Answer only yes or no.' Bypassed if FNOL_FILTER_ENABLED=false."),
    ]
    for title, desc in filters:
        story.append(Paragraph(f"<b>{title}:</b>  {desc}", styles["body"]))
        story.append(Spacer(1, 1.5*mm))

    story.append(Spacer(1, 2*mm))
    story.append(Paragraph("4.3 Policy Number Extraction (Regex Patterns)", styles["h2"]))
    story.append(Paragraph("Function: extract_policy_number()  in  backend/ingested_claims/service.py", styles["body"]))
    story.append(Spacer(1, 1*mm))

    patterns = [
        ("Keyword-anchored group 1", r"poli?cy\s*(?:number|no\.?|#|id|ref(?:erence)?)?\s*[:#\-=\s]\s*([A-Z0-9]{4,}(?:[-/][A-Z0-9]+)*)"),
        ("Claim number variant", r"claim\s*(?:number|no\.?|#|id|ref(?:erence)?)?\s*[:#\s]\s*([A-Z0-9]{4,}(?:[-/][A-Z0-9]+)*)"),
        ("Hash-prefix pattern", r"#([A-Z]{1,4}\d{5,})"),
        ("Structural – 2-letter prefix", r"\b([A-Z]{2}\d{6,})\b"),
        ("Structural – 3-letter prefix", r"\b([A-Z]{3}\d{6,})\b"),
        ("Structural – hyphenated", r"\b([A-Z]{2,4}-\d{5,})\b"),
        ("Fallback", "If none match → sender email address OR Message-ID display string used as placeholder"),
    ]
    for label, pattern in patterns:
        story.append(Paragraph(f"<b>{label}:</b>", styles["bullet"]))
        story.append(Paragraph(pattern, styles["code"]))
        story.append(Spacer(1, 1*mm))

    story.append(Spacer(1, 2*mm))
    story.append(Paragraph("4.4 MIME Email Parsing", styles["h2"]))
    mime_items = [
        "Body: _extract_body_text() walks MIME tree; first text/plain part preferred; fallback to text/html with _strip_html()",
        "Headers: _decode_header_value() handles RFC2047 MIME-encoded words; _format_address() normalises From/To",
        "Attachments: collected as {'filename', 'content_type', 'data': bytes}; saved to data/ingested-attachments/{claimId}/",
        "SendGrid path: extract_plain_body_from_rfc822() parses raw RFC822 field when text/html are empty",
        "Thread quoted text: _primary_message_text() strips quoted reply blocks; filters run only on fresh content",
    ]
    for item in mime_items:
        story.append(Paragraph(f"• {item}", styles["bullet"]))

    story.append(PageBreak())

    # ── Section 5: Extraction deep-dive ──────────────────
    story.append(section_banner("5. Claim Extraction – Deep Dive", C_BLUE, styles))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("5.1 Email Field Extraction (gpt-4o)", styles["h2"]))
    ext_kv = [
        ("Model", "gpt-4o  (env: OPENAI_MODEL, default gpt-4o)"),
        ("Temperature", "0.1"),
        ("Max tokens", "1500"),
        ("System prompt", "You are an expert insurance claims processor. Extract the following information from the FNOL email. Return ONLY valid JSON, no markdown."),
    ]
    story.append(kv_table(ext_kv, styles))
    story.append(Spacer(1, 2*mm))

    story.append(Paragraph("Extracted Fields:", styles["h3"]))
    fields = [
        ("policyNumber", "Policy number string from email body / subject"),
        ("claimantName", "Full name of the person filing the claim"),
        ("contactEmail", "Contact email of claimant"),
        ("contactPhone", "Contact phone number"),
        ("lossDate", "Date of loss/incident (ISO format preferred)"),
        ("lossType", "Enum: fire | water | auto | theft | liability | medical | other"),
        ("lossLocation", "Address or description of incident location"),
        ("description", "Full narrative description of the loss"),
        ("vehicleInfo", "Vehicle make / model / year / VIN (auto claims)"),
        ("propertyAddress", "Property address (home / commercial claims)"),
        ("estimatedAmount", "Claimant's estimated loss amount (numeric)"),
        ("_confidence", "Per-field confidence map (0.0 – 1.0) indicating extraction certainty"),
    ]
    fld_data = [[Paragraph(f"<b>{f}</b>", styles["cell"]), Paragraph(d, styles["cell"])]
                for f, d in fields]
    fld_t = Table(fld_data, colWidths=[38*mm, page_w - 38*mm])
    fld_t.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0,0), (-1,-1), [C_LBLUE, C_WHITE]),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#90CAF9")),
        ("TOPPADDING", (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING", (0,0),(-1,-1), 5),
        ("VALIGN", (0,0),(-1,-1), "TOP"),
    ]))
    story.append(fld_t)
    story.append(Spacer(1, 3*mm))

    story.append(Paragraph("5.2 Document / Attachment Extraction (gpt-4o)", styles["h2"]))
    doc_items = [
        "Document types classified: PoliceReport | RepairEstimate | Invoice | MedicalRecord | IncidentReport | DamagePhoto | Other",
        "_extract_schemas(): returns type-specific field hints per document type",
        "Model: gpt-4o, temp=0.1, max_tokens=1500 – same config as email extraction",
        "Returns: documentType, keyFields{}, confidence",
    ]
    for item in doc_items:
        story.append(Paragraph(f"• {item}", styles["bullet"]))
    story.append(Spacer(1, 2*mm))

    story.append(Paragraph("5.3 Vision / Image Analysis (gpt-4o multimodal)", styles["h2"]))
    vis_items = [
        "Input: image bytes base64-encoded → image_url data URL in message content list",
        "Model: gpt-4o (vision), max_tokens=2500",
        "Output fields: damage_severity, repair_cost_estimate, injury_indicators, detailed_summary, additional_observations",
        "Used for: damage photos, scanned documents with visual content",
    ]
    for item in vis_items:
        story.append(Paragraph(f"• {item}", styles["bullet"]))
    story.append(PageBreak())

    # ── Section 6: Policy Grounding ───────────────────────
    story.append(section_banner("6. Policy Grounding & Decision Pack", C_TEAL, styles))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("6.1 Local Database Grounding (primary path)", styles["h2"]))
    pg_items = [
        "Input: policyNumber (from LLM extraction) OR sender email",
        "find_customer_by_email(): email → customer record from database/local_data/customers.json",
        "find_policy_by_number(): policyNumber → policy record from database/local_data/policies.json",
        "Policy active check: status == 'active'  AND  expiry_date >= today",
        "Coverage lookup: database/policy_grounding_mapping.json maps lossType × policy_type → applicable clause IDs",
        "Policy types handled: AUTO / HOME / COMMERCIAL",
        "Per-clause confidence scoring: base=1.0; deductions for inactive policy, near-expiry (≤30 days), lossType mismatch",
        "Returns: { groundingClauses[], coverageConfirmed, policyStatus, policyHolderInfo }",
        "Policy status flags: POLICY-EXPIRED | POLICY-INACTIVE | POLICY-NOT-FOUND | CUSTOMER-NOT-FOUND",
    ]
    for item in pg_items:
        story.append(Paragraph(f"• {item}", styles["bullet"]))
    story.append(Spacer(1, 2*mm))

    story.append(Paragraph("6.2 ISO Clause Fallback (when local DB has no match)", styles["h2"]))
    iso_items = [
        "POLICY_CLAUSES: static library of ISO PAP / HO-3 / CGL-style clause texts in policy_clauses.py",
        "_infer_product_types(): policy number prefix (AC→auto, HO→home, CL→commercial)",
        "_infer_loss_types(): regex on lossType field + description narrative",
        "_compute_similarity(): keyword overlap between clause text and combined claim description",
        "Thresholds: CONFIDENCE_THRESHOLD_HIGH=0.8, MEDIUM=0.6; clauses with score < 0.6 excluded",
        "Max 6 clauses returned; sorted by descending similarity score",
    ]
    for item in iso_items:
        story.append(Paragraph(f"• {item}", styles["bullet"]))
    story.append(Spacer(1, 2*mm))

    story.append(Paragraph("6.3 Decision Pack Assembly", styles["h2"]))
    dp_items = [
        "build_decision_pack(claim_id, extraction_result) in backend/decision/service.py",
        "Merges all extracted fields; resolves policy via LLM result → DB lookup → fallback",
        "Runs local grounding first; if empty → ISO clause fallback",
        "Derives coverage_confirmed: True if grounding returned clauses + policy active",
        "Auto-reject flags: POLICY-EXPIRED / INACTIVE / NOT-FOUND / CUSTOMER-NOT-FOUND → triggers desk rejection in orchestrator",
        "Decision pack contains: claimId, timestamp, processingStatus, extractedFields (with _confidence), evidence[], documentAnalysis[], groundingClauses[], coverageConfirmed, policyHolderInfo, auditMetadata",
        "No LLM used in decision/grounding; entirely rule-based and DB-driven",
    ]
    for item in dp_items:
        story.append(Paragraph(f"• {item}", styles["bullet"]))
    story.append(PageBreak())

    # ── Section 7: Notifications ──────────────────────────
    story.append(section_banner("7. Notifications & FAQ Resolution", C_ORANGE, styles))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("7.1 Automated Email Notifications (claim_notification/service.py)", styles["h2"]))
    notif_rows = [
        ("FNOL Receipt Ack", "After IMAP/SendGrid ingest", "FNOL_RECEIPT_ACK_EMAIL_ENABLED", "'We received your claim, reference #{claimId}'"),
        ("Under Review", "After save_processed_claim() (non-rejected)", "CLAIM_UNDER_REVIEW_EMAIL_ENABLED", "Claim details, policy holder info, next steps"),
        ("Desk Rejection", "Orchestrator detects expired/invalid policy", "DESK_REJECTION_EMAIL_ENABLED", "Rejection reason, policy status, appeal instructions"),
    ]
    notif_h = [Paragraph(h, styles["cell_hdr"]) for h in ["Notification", "Trigger", "Toggle Env Var", "Content"]]
    notif_data = [notif_h] + [[Paragraph(v, styles["cell"]) for v in r] for r in notif_rows]
    notif_t = Table(notif_data, colWidths=[35*mm, 42*mm, 50*mm, page_w - 127*mm])
    notif_t.setStyle(TableStyle([
        ("BACKGROUND", (0,0),(-1,0), C_ORANGE),
        ("ROWBACKGROUNDS", (0,1),(-1,-1), [C_LORANGE, C_WHITE]),
        ("GRID", (0,0),(-1,-1), 0.3, colors.HexColor("#FFCC02")),
        ("TOPPADDING", (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING", (0,0),(-1,-1), 4),
        ("VALIGN", (0,0),(-1,-1), "TOP"),
    ]))
    story.append(notif_t)
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph("All notifications use smtplib SMTP with In-Reply-To / References headers to thread into original email.", styles["body"]))
    story.append(Spacer(1, 3*mm))

    story.append(Paragraph("7.2 FAQ Resolution (faq_resolution/service.py)", styles["h2"]))
    faq_items = [
        "Triggered: email classified as non-FNOL by email_ingestion or directly via source=faq inbox",
        "Dedup: check faq-answered-ids.json; skip if message_id already answered",
        "_is_faq_query(): regex keyword patterns first; then gpt-4o-mini 'FAQ or CLAIM?' classification",
        "Customer data intent: _classify_customer_data_intent() → detect 'what is my deductible / policy limit / status' → answer from local JSON DB",
        "FAQ answer: _find_faq_answer() → gpt-4o-mini receives full data/FAQ.csv; prompted to pick 'Q1', 'Q2', … index; extract answer",
        "Keyword fallback: if no API key → overlap score between question text and FAQ rows",
        "_send_faq_response_email(): SMTP threaded reply (In-Reply-To / References set)",
        "Persist: message_id stored to faq-answered-ids.json to prevent duplicate replies",
    ]
    for item in faq_items:
        story.append(Paragraph(f"• {item}", styles["bullet"]))
    story.append(PageBreak())

    # ── Section 8: FastAPI + Frontend ────────────────────
    story.append(section_banner("8. FastAPI API Layer & Frontend", C_NAVY, styles))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("8.1 FastAPI REST Endpoints", styles["h2"]))
    ep_h = [Paragraph(h, styles["cell_hdr"]) for h in ["Method", "Path", "Purpose"]]
    ep_data = [ep_h] + [[Paragraph(e[0], styles["cell"]), Paragraph(e[1], styles["cell"]), Paragraph(e[2], styles["cell"])] for e in ENDPOINTS]
    ep_t = Table(ep_data, colWidths=[16*mm, 58*mm, page_w - 74*mm])
    ep_style = [
        ("BACKGROUND", (0,0),(-1,0), C_NAVY),
        ("GRID", (0,0),(-1,-1), 0.3, colors.HexColor("#9FA8DA")),
        ("TOPPADDING", (0,0),(-1,-1), 4),
        ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ("LEFTPADDING", (0,0),(-1,-1), 4),
        ("VALIGN", (0,0),(-1,-1), "TOP"),
    ]
    for i, ep in enumerate(ENDPOINTS, 1):
        bg = C_LGREEN if ep[0] == "GET" else C_LORANGE
        ep_style.append(("BACKGROUND", (0,i), (-1,i), bg))
    ep_t.setStyle(TableStyle(ep_style))
    story.append(ep_t)
    story.append(Spacer(1, 3*mm))

    story.append(Paragraph("8.2 Frontend Components (Next.js)", styles["h2"]))
    fe_rows = [
        ("app/page.tsx", "Stage router: home → review → decision → dashboard; auth gate; FAQ page toggle"),
        ("HomePage.tsx", "Ingested claim inbox; IMAP sync via POST /api/sync-inbox; process claim via POST /api/process-claim"),
        ("ReviewPage.tsx", "Displays extracted fields + evidence list for adjuster review"),
        ("DecisionPage.tsx", "Shows decision pack: policy grounding, coverage confirmed, policy holder info"),
        ("DashboardPage.tsx", "KPI charts via GET /api/dashboard/kpis; Recharts visualisations"),
        ("FAQPage.tsx", "FAQ-sourced inbox (source=faq query param to ingested-claims endpoint)"),
        ("Header.tsx", "Navigation bar between pages"),
        ("ClaimSummaryBar.tsx", "Persistent claim summary strip across all stages"),
        ("lib/backend.ts", "runPython() helper for legacy webhook spawning"),
        ("lib/agents/orchestrator.ts", "LangGraph-style client-side orchestrator (present but NOT used in main claim flow – FastAPI path is active)"),
        ("app/api/webhooks/sendgrid-inbound/route.ts", "Receives SendGrid inbound-parse POST; forwards to FastAPI ingested-claims endpoint"),
    ]
    fe_data = [[Paragraph(f"<b>{f}</b>", styles["cell"]), Paragraph(r, styles["cell"])] for f, r in fe_rows]
    fe_t = Table(fe_data, colWidths=[60*mm, page_w - 60*mm])
    fe_t.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0,0),(-1,-1), [colors.HexColor("#E8EAF6"), C_WHITE]),
        ("GRID", (0,0),(-1,-1), 0.3, colors.HexColor("#9FA8DA")),
        ("TOPPADDING", (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING", (0,0),(-1,-1), 5),
        ("VALIGN", (0,0),(-1,-1), "TOP"),
    ]))
    story.append(fe_t)
    story.append(PageBreak())

    # ── Section 9: Data stores + Env vars ────────────────
    story.append(section_banner("9. Data Stores & Environment Variables", C_TEAL, styles))
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("9.1 Data Storage (JSON / CSV – no SQL at runtime)", styles["h2"]))
    ds_data = [[Paragraph(f"<b>{p}</b>", styles["cell"]),
                Paragraph(r, styles["cell"]),
                Paragraph(d, styles["cell"])] for p, r, d in DATA_STORES]
    ds_t = Table(ds_data, colWidths=[58*mm, 28*mm, page_w - 86*mm])
    ds_t.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0,0),(-1,-1), [C_LTEAL, C_WHITE]),
        ("GRID", (0,0),(-1,-1), 0.3, colors.HexColor("#80CBC4")),
        ("TOPPADDING", (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING", (0,0),(-1,-1), 5),
        ("VALIGN", (0,0),(-1,-1), "TOP"),
    ]))
    story.append(ds_t)
    story.append(Spacer(1, 4*mm))

    story.append(Paragraph("9.2 Environment Variables", styles["h2"]))
    ev_data = [[Paragraph(f"<b>{v}</b>", styles["cell"]),
                Paragraph(c, styles["cell"]),
                Paragraph(d, styles["cell"])] for v, c, d in ENV_VARS]
    ev_t = Table(ev_data, colWidths=[55*mm, 24*mm, page_w - 79*mm])
    ev_t.setStyle(TableStyle([
        ("ROWBACKGROUNDS", (0,0),(-1,-1), [colors.HexColor("#E8EAF6"), C_WHITE]),
        ("GRID", (0,0),(-1,-1), 0.3, colors.HexColor("#9FA8DA")),
        ("TOPPADDING", (0,0),(-1,-1), 3),
        ("BOTTOMPADDING", (0,0),(-1,-1), 3),
        ("LEFTPADDING", (0,0),(-1,-1), 5),
        ("VALIGN", (0,0),(-1,-1), "TOP"),
    ]))
    story.append(ev_t)

    doc.build(story)
    print(f"PDF saved → {out_path}")


# ─────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────

if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    os.makedirs(base, exist_ok=True)

    xlsx_path = os.path.join(base, "Autonomous_Claims_Orchestrator_Workflow.xlsx")
    pdf_path  = os.path.join(base, "Autonomous_Claims_Orchestrator_Workflow.pdf")

    build_xlsx(xlsx_path)
    build_pdf(pdf_path)
    print("Done.")
